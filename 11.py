# Работа с exel
import openpyxl
# Открываем файл exel, где в sample - имя файла, причём, должен быть в папке с пайтон файлом, а то не найдёт
wb = openpyxl.reader.excel.load_workbook(filename="sample.xlsx")
# Выводим на экран листы этой командой [то есть будет вывод в виде списка]
print(wb.sheetnames)
# Выбираем первый лист активным для работы с ним, путём ввода индекса
wb.active = 0
# Делаем ячейки активными
sheet = wb.active
# Для вывода значения(value) определённой ячейки
print(sheet['A1'].value)
# Цикл для прогона всего списка (не забудь преобразовать i  в строку)
for i in range(1, 12):
	print(sheet['A' + str(i)].value, sheet['B' + str(i)].value, sheet['C' + str(i)].value, sep='+')
# Чтобы вместо формул при выводе значений появлялся ответ
wb = openpyxl.reader.excel.load_workbook(filename="sample.xlsx", data_only=True)
wb.active = 1
sheet = wb.active
for i in range(1, 12):
	print(sheet['A' + str(i)].value, sheet['B' + str(i)].value, sheet['C' + str(i)].value, sep='-')
# Пока файл открыт, выдаст ошибку, перед сохранением закрой (можно использовать другое имя файла, тогда всё уйдёт туда)
wb.save('sample.xlsx')
#  Вставил подсказку при импорте модуля
'''
from openpyxl import Workbook
wb = Workbook()

# захватите активный рабочий лист
ws = wb.active

# Данные могут быть назначены непосредственно ячейкам
ws['A1'] = 42

# Строки также могут быть добавлены
ws.append([1, 2, 3])

# Типы Python будут автоматически преобразованы
import datetime
ws['A2'] = datetime.datetime.now()

# Сохраните файл
wb.save("sample.xlsx")
'''